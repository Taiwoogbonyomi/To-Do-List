import datetime
import openpyxl
from openpyxl import Workbook

SPREADSHEET_FILE = "tasks.xlsx"

def load_tasks():
    """
    Load tasks from the spreadsheet.
    """
    tasks = []
    try:
        workbook = openpyxl.load_workbook(SPREADSHEET_FILE)
        sheet = workbook.active
        tasks = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            task = {
                "description": row[0],
                "due_date": row[1],
                "due_time": row[2]
            }
            tasks.append(task)

    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Description", "Due Date", "Due Time"])
        workbook.save(SPREADSHEET_FILE)

def save_tasks():
    """Save tasks to the spreadsheet."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Description", "Due Date", "Due Time"])
    for task in tasks:
        sheet.append([task["description"], task["due_date"], task["due_time"]])
    workbook.save(SPREADSHEET_FILE)

#This is an empty list to store tasks
tasks = []

def display_instructions():
    """
    Function to display the to_do list instructions
    """
    print("""
    ===== To-Do List Instructions =====
    Welcome to the To-Do list application!
    You can manage your tasks with the following options:
    1. Display to-do list : View all tasks currently in the list.
    2. Add a task : Enter a new task to add to your to-do list.
    3. Remove a task : Choose a task by its index to remove the task from your list.
    4. Quit : Exit the application.
    ===================================== 
    
    """)

#Function to display the to-do list
def display_tasks():
        if not tasks:
            print("\U0001F4ED Your to-do list is empty.")
        else:
            print("\U0001F4CB Your To-Do List.")
            for index,task in enumerate(tasks, start=1):
                print(f"{index}. {task['description']} - Due: {task['due_date']} {task['due_time']}")


#Function to add task to the to-do list
def add_task():
        description = input("Enter task description: ") 
        due_date = input("Enter due date (YYYY-MM-DD): ")
        due_time = input("Enter due-time (HH:MM): ")
        try:
             datetime.datetime.strptime(due_date, "%Y-%m-%d")
             datetime.datetime.strptime(due_time, "%H:%M")       
        except ValueError:
             print("\U0000274C Invalid date or time format. Please try again.")
             return

        task = {
            "description": description,
            "due_date": due_date,
            "due_time": due_time
        }
        
        tasks.append(task)
        print(f"\U00002705 Task '{description}' added to your to-do list")
        save_tasks()   
    

#Function to remove a task from the to-do list
def remove_task():
    display_tasks()
    if tasks:
        try:
            index = int(input("Enter the number of the task to remove: ")) - 1
            if 0 <= index < len(tasks):
                removed_task = tasks[index]
                confirm = input(f"Are you sure you want to remove the task '{removed_task['description']}'? (yes/no): ").lower()
                if confirm == 'yes':
                    tasks.pop(index)
                    print(f"\U0001F4ED Task '{removed_task['description']}' removed successfully.")
                    save_tasks()
                else:
                    print("\U0000274C Task removal cancelled.")
            else:
                print("\U0000274C Invalid input")
        except ValueError:
            print("\U0000274C Invalid input. Please enter a number.")
    else:
        print("\U0001F4ED Your to-do list is already empty")            


def main():
        
        """
        Main function to run the to-do list application
        """
        load_tasks()
        display_instructions()
        print("\U0001F4CB Welcome to the to-do list app \U0001F4CB")
        while True:
            print("\n === To-Do List Menu===")
            print("Please select one of the following options")
            print("------------------------------------------")
            print("1. Display to-do list")
            print("2. Add a task")
            print("3. Remove a task")
            print("4. Quit")

            choice = input("Enter your choice (1-4): ")
            
            if (choice == "1"):
                display_tasks()
            elif (choice == "2"):
                add_task()
            elif (choice == "3"):
                remove_task()
            elif (choice == "4"):
                print("Quit application...")
                print(" Goodbye \U0001F44B \U0001F44B ")
                break
            else:
                print("\U0000274C Invalid choice, Please enter a number from 1 to 4")
                
            

# Run the main function to start the application
if __name__ == "__main__":
    main()


